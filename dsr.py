import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C_DARK_BG="0A0A0B"; C_PANEL="111113"; C_CARD="18181B"; C_GOLD="F5C842"
C_TEAL="2DD4BF"; C_WHITE="F4F4F5"; C_GRAY="A1A1AA"; C_RED="F87171"
C_PURPLE="A78BFA"; C_GREEN="34D399"; C_ORANGE="FB923C"

def bg(h): return PatternFill("solid", fgColor=h)
def ft(h, size=11, bold=False, italic=False): return Font(name="Arial", color=h, size=size, bold=bold, italic=italic)
def al(h="left", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def tb(color="3A3A42"):
    s=Side(style="thin",color=color); return Border(left=s,right=s,top=s,bottom=s)

def sc(ws,row,col,val,fill=None,font=None,alignment=None,border=None,nf=None):
    c=ws.cell(row=row,column=col,value=val)
    if fill: c.fill=fill
    if font: c.font=font
    if alignment: c.alignment=alignment
    if border: c.border=border
    if nf: c.number_format=nf
    return c

def hdr(ws,row,cols,bgc=C_GOLD,fgc=C_DARK_BG,size=10):
    for col,val in cols:
        sc(ws,row,col,val,fill=bg(bgc),font=ft(fgc,size=size,bold=True),alignment=al("center"),border=tb(bgc))

def drow(ws,row,cols,bgc=C_CARD,fgc=C_WHITE,alt=False):
    fc="1F1F24" if alt else bgc
    for col,val in cols:
        sc(ws,row,col,val,fill=bg(fc),font=ft(fgc,size=10),alignment=al("center"),border=tb("2A2A30"))

def mth(ws,sr,sc2,er,ec,text,bgc=C_PANEL,fgc=C_GOLD,size=13,bold=True):
    ws.merge_cells(start_row=sr,start_column=sc2,end_row=er,end_column=ec)
    c=ws.cell(row=sr,column=sc2,value=text)
    c.fill=bg(bgc); c.font=ft(fgc,size=size,bold=bold); c.alignment=al("center")

def fill_bg(ws,max_row=80,max_col=25):
    for row in ws.iter_rows(min_row=1,max_row=max_row,min_col=1,max_col=max_col):
        for cell in row: cell.fill=bg(C_DARK_BG)

# ── Load data ──────────────────────────────────────────────────────────────────
df = pd.read_excel("C:/Users/Tahir General/Downloads/Detailed Sales Report.xlsx")
df['Date'] = pd.to_datetime(df['Date'], format='%m/%d/%Y-%I:%M %p', errors='coerce')
df['sale_date'] = df['Date'].dt.date
df['hour'] = df['Date'].dt.hour
df['day_of_week'] = df['Date'].dt.day_name()

df_clean = df[(df['Date'] >= '2026-04-11') & (df['Name'] != 'Discount') & (df['Quantity Sold'] > 0)].copy()
def time_bucket(h):
    if h < 11: return 'Morning(<11)'
    elif h < 13: return 'Midday(11-1)'
    elif h < 18: return 'Afternoon(1-6)'
    else: return 'Evening(6-9)'
df_clean['time_bucket'] = df_clean['hour'].apply(time_bucket)
bins=[0,5000,10000,20000,30000,50000,100000,float('inf')]
labels=['<5k','5-10k','10-20k','20-30k','30-50k','50-100k','>100k']
df_clean['price_bucket'] = pd.cut(df_clean['Selling Price'],bins=bins,labels=labels)

total_rev = df_clean['Subtotal'].sum()
total_txn = df_clean['Sale Id'].nunique()
total_units = df_clean['Quantity Sold'].sum()
num_days = df_clean['sale_date'].nunique()

cat_df = df_clean.groupby('Category').agg(Revenue=('Subtotal','sum'),Units=('Quantity Sold','sum'),Transactions=('Sale Id','nunique')).sort_values('Revenue',ascending=False).reset_index()
cat_df['AvgPrice'] = cat_df['Revenue']/cat_df['Units']

dow_order=['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
dow_df = df_clean.groupby('day_of_week').agg(Revenue=('Subtotal','sum'),Transactions=('Sale Id','nunique'),Units=('Quantity Sold','sum')).reindex(dow_order).fillna(0).reset_index()
total_week_rev = dow_df['Revenue'].sum()

top_df = df_clean.groupby(['Name','Category']).agg(Revenue=('Subtotal','sum'),Units=('Quantity Sold','sum'),Transactions=('Sale Id','nunique')).sort_values('Revenue',ascending=False).reset_index()
top_df['AvgPrice'] = top_df['Revenue']/top_df['Units']
top_df['RevPerTxn'] = top_df['Revenue']/top_df['Transactions']

tb_df = df_clean.groupby('time_bucket').agg(Transactions=('Sale Id','nunique'),Revenue=('Subtotal','sum')).reset_index()
tb_total = tb_df['Revenue'].sum()

pb_df = df_clean.groupby('price_bucket',observed=True).agg(LineItems=('Name','count'),Units=('Quantity Sold','sum'),Revenue=('Subtotal','sum')).reset_index()
pb_df['PctRevenue'] = pb_df['Revenue']/pb_df['Revenue'].sum()

daily_df = df_clean.groupby('sale_date').agg(Revenue=('Subtotal','sum'),Transactions=('Sale Id','nunique'),Units=('Quantity Sold','sum')).reset_index()

best_day_by_cat = df_clean.groupby(['Category','day_of_week'])['Subtotal'].sum().reset_index()
best_day = best_day_by_cat.loc[best_day_by_cat.groupby('Category')['Subtotal'].idxmax()].reset_index(drop=True)

top_cust = df_clean.groupby('Sold To').agg(Revenue=('Subtotal','sum'),Transactions=('Sale Id','nunique'),Units=('Quantity Sold','sum')).sort_values('Revenue',ascending=False).head(15).reset_index()

period_avg = total_rev/num_days

# ── WORKBOOK ───────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ═══════════════════════ SHEET 1: EXECUTIVE DASHBOARD ════════════════════════
ws1 = wb.create_sheet("Executive Dashboard")
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = C_GOLD
fill_bg(ws1, 80, 25)
for i in range(1,26): ws1.column_dimensions[get_column_letter(i)].width = 13

# Title
ws1.row_dimensions[2].height = 42; ws1.row_dimensions[3].height = 20
mth(ws1,2,1,2,25,"TEXTILE BUSINESS — SALES INTELLIGENCE REPORT",C_DARK_BG,C_GOLD,22)
mth(ws1,3,1,3,25,"Period: 11 Apr – 26 Apr 2026  |  POS Analytics  •  16 Days",C_DARK_BG,C_GRAY,11)

# KPI row
kpis = [
    (f"₦{total_rev:,.0f}", "TOTAL REVENUE (16 DAYS)", C_GOLD, 2),
    (f"{total_txn:,}", "TOTAL TRANSACTIONS", C_TEAL, 8),
    (f"₦{total_rev/num_days:,.0f}", "AVG DAILY REVENUE", C_PURPLE, 14),
    (f"₦{total_rev/total_txn:,.0f}", "AVG BASKET VALUE", C_GREEN, 20),
]
for val, label, color, sc2 in kpis:
    ws1.row_dimensions[5].height = 8; ws1.row_dimensions[6].height = 38; ws1.row_dimensions[7].height = 22
    mth(ws1,6,sc2,6,sc2+4,val,C_PANEL,color,18)
    mth(ws1,7,sc2,7,sc2+4,label,C_PANEL,C_GRAY,9)

# Tables header
ws1.row_dimensions[9].height = 12
for sr,ec,text,color in [(10,6,"  📅  REVENUE BY DAY OF WEEK",C_GOLD),(10,13,"  🧵  REVENUE BY FABRIC CATEGORY",C_GOLD),(10,20,"  📆  DAILY REVENUE TRACKER",C_GOLD)]:
    start = {6:2, 13:8, 20:14}[ec]
    mth(ws1,10,start,10,ec,text,C_PANEL,color,11)

hdr(ws1,11,[(2,"Day"),(3,"Revenue (₦)"),(4,"Transactions"),(5,"Units"),(6,"% Week")])
for i,(_,r) in enumerate(dow_df.iterrows()):
    row=12+i; alt=i%2==1
    pct=r['Revenue']/total_week_rev if total_week_rev else 0
    drow(ws1,row,[(2,r['day_of_week']),(3,int(r['Revenue'])),(4,int(r['Transactions'])),(5,int(r['Units'])),(6,f"{pct:.1%}")],alt=alt)
    ws1.cell(row=row,column=3).number_format='#,##0'

hdr(ws1,11,[(8,"Category"),(9,"Revenue (₦)"),(10,"Units"),(11,"Avg Price"),(12,"% Share")])
for i,(_,r) in enumerate(cat_df.iterrows()):
    row=12+i; alt=i%2==1
    drow(ws1,row,[(8,r['Category']),(9,int(r['Revenue'])),(10,int(r['Units'])),(11,int(r['AvgPrice'])),(12,f"{r['Revenue']/total_rev:.1%}")],alt=alt)
    ws1.cell(row=row,column=9).number_format='#,##0'; ws1.cell(row=row,column=11).number_format='#,##0'

hdr(ws1,11,[(14,"Date"),(15,"Revenue (₦)"),(16,"Txns"),(17,"Units")])
for i,(_,r) in enumerate(daily_df.iterrows()):
    row=12+i; alt=i%2==1
    drow(ws1,row,[(14,str(r['sale_date'])),(15,int(r['Revenue'])),(16,int(r['Transactions'])),(17,int(r['Units']))],alt=alt)
    ws1.cell(row=row,column=15).number_format='#,##0'

# Insights
ins_row=23
mth(ws1,ins_row+1,2,ins_row+1,22,"  💡  KEY PERFORMANCE INSIGHTS — 16 DAYS (APR 11–26, 2026)","1A1A1F",C_GOLD,12)
insights=[
    f"📈  TOTAL REVENUE: ₦{total_rev:,.0f} across {num_days} trading days = ₦{total_rev/num_days:,.0f}/day average",
    f"🏆  SATURDAY is peak day: ₦{int(dow_df[dow_df['day_of_week']=='Saturday']['Revenue'].values[0]):,} = {dow_df[dow_df['day_of_week']=='Saturday']['Revenue'].values[0]/total_week_rev:.0%} of weekly revenue",
    f"🧵  ATAMPA dominates at {cat_df[cat_df['Category']=='ATAMPA']['Revenue'].values[0]/total_rev:.0%} of revenue — stock deeply",
    f"⏰  EVENING (6-9pm) drives ₦{int(tb_df[tb_df['time_bucket']=='Evening(6-9)']['Revenue'].values[0]):,} — key shift from previous period",
    f"👑  FAHAD TAHIR is #1 customer at ₦{int(top_cust.iloc[0]['Revenue']):,} across {int(top_cust.iloc[0]['Transactions'])} transactions",
    f"🚀  MEDIUM SUPER is #1 product with ₦{int(top_df.iloc[0]['Revenue']):,} from {int(top_df.iloc[0]['Units'])} units — new leader",
    f"📊  ₦10k–₦20k price band = {pb_df[pb_df['price_bucket']=='10-20k']['PctRevenue'].values[0]:.0%} of revenue — sweet spot, stock this deeply",
    f"💡  APR 25 was best single day: ₦{int(daily_df.nlargest(1,'Revenue')['Revenue'].values[0]):,} — replicate the conditions",
]
for i,ins in enumerate(insights):
    row=ins_row+2+i; ws1.row_dimensions[row].height=18
    mth(ws1,row,2,row,22,ins,"1A1A1F" if i%2==0 else C_CARD,C_WHITE if i%2==0 else C_GRAY,10,False)

# ═══════════════════════ SHEET 2: TIMING INTELLIGENCE ════════════════════════
ws2 = wb.create_sheet("Timing Intelligence")
ws2.sheet_view.showGridLines=False; ws2.sheet_properties.tabColor=C_TEAL
fill_bg(ws2,80,20)
for i in range(1,20): ws2.column_dimensions[get_column_letter(i)].width=18

mth(ws2,1,1,1,12,"POSTING TIMING INTELLIGENCE",C_DARK_BG,C_GOLD,20)
mth(ws2,2,1,2,12,"Optimal selling windows  •  Apr 11–26 2026",C_DARK_BG,C_GRAY,11)

mth(ws2,4,1,4,4,"  🕐  SALES BY TIME OF DAY",C_PANEL,C_TEAL,12)
hdr(ws2,5,[(1,"Time Bucket"),(2,"Transactions"),(3,"Revenue (₦)"),(4,"% Revenue")],bgc=C_TEAL,fgc=C_DARK_BG)
for i,(_,r) in enumerate(tb_df.iterrows()):
    row=6+i; alt=i%2==1
    drow(ws2,row,[(1,r['time_bucket']),(2,int(r['Transactions'])),(3,int(r['Revenue'])),(4,f"{r['Revenue']/tb_total:.1%}")],alt=alt)
    ws2.cell(row=row,column=3).number_format='#,##0'

mth(ws2,4,6,4,9,"  🏆  BEST DAY BY CATEGORY",C_PANEL,C_TEAL,12)
hdr(ws2,5,[(6,"Category"),(7,"Best Day"),(8,"Revenue (₦)")],bgc=C_TEAL,fgc=C_DARK_BG)
for i,(_,r) in enumerate(best_day.iterrows()):
    row=6+i; alt=i%2==1
    drow(ws2,row,[(6,r['Category']),(7,r['day_of_week']),(8,int(r['Subtotal']))],alt=alt)
    ws2.cell(row=row,column=8).number_format='#,##0'

mth(ws2,14,1,14,9,"  📅  DAY-BY-DAY DETAILED BREAKDOWN",C_PANEL,C_TEAL,12)
hdr(ws2,15,[(1,"Day"),(2,"Date"),(3,"Revenue (₦)"),(4,"Transactions"),(5,"Units"),(6,"Avg Basket"),(7,"vs Avg")],bgc=C_TEAL,fgc=C_DARK_BG)
for i,(_,r) in enumerate(daily_df.iterrows()):
    row=16+i; alt=i%2==1
    vs=((r['Revenue']-period_avg)/period_avg)
    dow_name=pd.Timestamp(str(r['sale_date'])).strftime('%A')
    fg=C_GREEN if vs>0 else C_RED if vs<-0.2 else C_WHITE
    drow(ws2,row,[(1,dow_name),(2,str(r['sale_date'])),(3,int(r['Revenue'])),(4,int(r['Transactions'])),(5,int(r['Units'])),(6,int(r['Revenue']/r['Transactions'])),(7,f"{vs:+.1%}")],alt=alt,fgc=fg)
    ws2.cell(row=row,column=3).number_format='#,##0'
    ws2.cell(row=row,column=6).number_format='#,##0'

ins_r=33
mth(ws2,ins_r,1,ins_r,9,"💡 TIMING INSIGHTS — Expert Commercial Analysis","1A1A1F",C_GOLD,12)
sat_rev=int(dow_df[dow_df['day_of_week']=='Saturday']['Revenue'].values[0])
eve_rev=int(tb_df[tb_df['time_bucket']=='Evening(6-9)']['Revenue'].values[0])
wed_rev=int(dow_df[dow_df['day_of_week']=='Wednesday']['Revenue'].values[0])
best_day_row=daily_df.nlargest(1,'Revenue').iloc[0]
timing_ins=[
    f"• SATURDAY is peak day: ₦{sat_rev:,} = {sat_rev/total_week_rev:.0%} of weekly revenue. Post major arrivals FRIDAY 8–10pm.",
    f"• EVENING (6–9pm) drives ₦{eve_rev:,} — buyers browse socials after work. NEW pattern vs previous period.",
    f"• WEDNESDAY surprise: ₦{wed_rev:,} — investigate what drove this. Post premium ATAMPA Tuesday evening.",
    f"• APR 25 was best day: ₦{int(best_day_row['Revenue']):,} with {int(best_day_row['Transactions'])} transactions — study promotions used.",
    f"• RECOMMENDATION: Post ATAMPA Friday evening, LACE Saturday morning, MATERIAL Wednesday.",
]
for i,ins in enumerate(timing_ins):
    row=ins_r+1+i; ws2.row_dimensions[row].height=20
    mth(ws2,row,1,row,9,ins,"1A1A1F" if i%2==0 else C_CARD,C_WHITE,10,False)

# ═══════════════════════ SHEET 3: PRODUCT PERFORMANCE ════════════════════════
ws3 = wb.create_sheet("Product Performance")
ws3.sheet_view.showGridLines=False; ws3.sheet_properties.tabColor=C_PURPLE
fill_bg(ws3,80,12)
for i,w in enumerate([4,28,14,14,14,14,18,18],1): ws3.column_dimensions[get_column_letter(i)].width=w

mth(ws3,1,1,1,8,"PRODUCT PERFORMANCE ANALYTICS",C_DARK_BG,C_GOLD,20)
mth(ws3,2,1,2,8,"Revenue, velocity, and margin by product  •  Apr 11–26 2026",C_DARK_BG,C_GRAY,11)
mth(ws3,4,1,4,8,"  🥇  TOP 25 PRODUCTS BY REVENUE",C_PANEL,C_PURPLE,12)

hdr(ws3,5,[(1,"#"),(2,"Product"),(3,"Category"),(4,"Revenue (₦)"),(5,"Units"),(6,"Transactions"),(7,"Avg Price (₦)"),(8,"Rev/Txn (₦)")],bgc=C_PURPLE)
for i,(_,r) in enumerate(top_df.head(25).iterrows()):
    row=6+i; alt=i%2==1
    drow(ws3,row,[(1,i+1),(2,r['Name']),(3,r['Category']),(4,int(r['Revenue'])),(5,int(r['Units'])),(6,int(r['Transactions'])),(7,int(r['AvgPrice'])),(8,int(r['RevPerTxn']))],alt=alt)
    ws3.cell(row=row,column=4).font=ft(C_GOLD,10,bold=True)
    for col in [4,7,8]: ws3.cell(row=row,column=col).number_format='#,##0'

mth(ws3,32,1,32,8,"  🆕  NEW PRODUCTS THIS PERIOD (not in Period 1 top 20)",C_PANEL,C_GREEN,12)
new_names=['Elizabeth Embellished','ELIZABETH EMBELLISHED','Vibes Metallic Cotton Satin','VIBES METALLIC COTTON SATIN',
           'Hollantex Mix pattern','HOLLANTEX MIX PATTERN','Julius Embellished','JULIUS EMBELLISHED',
           'Victoire Magic Princess','VICTOIRE MAGIC PRINCESS','New Julius Holland','NEW JULIUS HOLLAND',
           'JULIUS HOLLAND PLAIN','Julius Holland Plain','MEDIUM SUPER','Medium Super']
new_df = top_df[top_df['Name'].isin(new_names)].head(10)
hdr(ws3,33,[(1,"#"),(2,"Product"),(3,"Category"),(4,"Revenue (₦)"),(5,"Units"),(6,"Transactions"),(7,"Avg Price (₦)")],bgc=C_GREEN,fgc=C_DARK_BG)
for i,(_,r) in enumerate(new_df.iterrows()):
    row=34+i; alt=i%2==1
    drow(ws3,row,[(1,i+1),(2,r['Name']),(3,r['Category']),(4,int(r['Revenue'])),(5,int(r['Units'])),(6,int(r['Transactions'])),(7,int(r['AvgPrice']))],alt=alt)
    for col in [4,7]: ws3.cell(row=row,column=col).number_format='#,##0'

# ═══════════════════════ SHEET 4: PRICE SENSITIVITY ════════════════════════
ws4 = wb.create_sheet("Price Sensitivity")
ws4.sheet_view.showGridLines=False; ws4.sheet_properties.tabColor=C_ORANGE
fill_bg(ws4,60,12)
for i in range(1,12): ws4.column_dimensions[get_column_letter(i)].width=16

mth(ws4,1,1,1,9,"PRICE SENSITIVITY ANALYSIS",C_DARK_BG,C_GOLD,20)
mth(ws4,2,1,2,9,"Optimal price bands — volume vs value tradeoff  •  Apr 11–26 2026",C_DARK_BG,C_GRAY,11)
mth(ws4,4,1,4,6,"  📊  SALES BY PRICE BUCKET",C_PANEL,C_ORANGE,12)

hdr(ws4,5,[(1,"Price Range"),(2,"Line Items"),(3,"Units Sold"),(4,"Revenue (₦)"),(5,"% of Revenue")],bgc=C_ORANGE,fgc=C_DARK_BG)
for i,(_,r) in enumerate(pb_df.iterrows()):
    row=6+i; alt=i%2==1
    drow(ws4,row,[(1,str(r['price_bucket'])),(2,int(r['LineItems'])),(3,int(r['Units'])),(4,int(r['Revenue'])),(5,f"{r['PctRevenue']:.1%}")],alt=alt)
    ws4.cell(row=row,column=4).number_format='#,##0'
    if str(r['price_bucket'])=='10-20k':
        for col in range(1,6): ws4.cell(row=row,column=col).font=ft(C_GOLD,10,bold=True)

mth(ws4,15,1,15,9,"  💡  PRICE STRATEGY RECOMMENDATIONS",C_PANEL,C_ORANGE,12)
strategies=[
    ("Volume Sweet Spot","₦10,000–₦20,000",f"{pb_df[pb_df['price_bucket']=='10-20k']['PctRevenue'].values[0]:.0%} of revenue — highest velocity. Stock deeply: MEDIUM SUPER, EMBELLISHED MDT, HOLLANDAD."),
    ("Revenue Sweet Spot","₦20,000–₦50,000",f"₦{int(pb_df[pb_df['price_bucket'].isin(['20-30k','30-50k'])]['Revenue'].sum()):,} combined. VIBES METALLIC & ABS MAITABARMA excel here."),
    ("Ultra Premium","> ₦100,000",f"₦{int(pb_df[pb_df['price_bucket']=='>100k']['Revenue'].values[0]):,} from {int(pb_df[pb_df['price_bucket']=='>100k']['LineItems'].values[0])} line items. Target VIP buyers only."),
    ("New Product Growth","₦30,000–₦50,000","ELIZABETH EMBELLISHED & VIBES METALLIC are new entrants excelling here. Increase stock immediately."),
    ("Bulk Incentive","₦5,000–₦10,000","HOLLANDAD and CHIGANVY INDIA sell in bulk. Offer 5% off at 10+ units."),
    ("Slow-Mover Alert","> ₦50,000","If >7 days unsold: repost with 5–10% discount or bundle with mid-range item."),
]
ws4.column_dimensions['D'].width=50
for i in range(1,10): ws4.column_dimensions[get_column_letter(i)].width=16
ws4.column_dimensions['D'].width=50
hdr(ws4,16,[(1,"Strategy"),(2,"Price Range"),(4,"Recommendation")],bgc=C_ORANGE,fgc=C_DARK_BG)
ws4.merge_cells('D16:I16')
ws4.cell(row=16,column=4,value="Recommendation").fill=bg(C_ORANGE)
ws4.cell(row=16,column=4).font=ft(C_DARK_BG,10,bold=True)
for i,(strat,price,rec) in enumerate(strategies):
    row=17+i
    ws4.merge_cells(f'D{row}:I{row}')
    drow(ws4,row,[(1,strat),(2,price),(4,rec)],alt=i%2==1)
    ws4.cell(row=row,column=1).font=ft(C_GOLD,10,bold=True)
    ws4.cell(row=row,column=4).alignment=al("left",wrap=True)

# ═══════════════════════ SHEET 5: CUSTOMER INTELLIGENCE ════════════════════════
ws5 = wb.create_sheet("Customer Intelligence")
ws5.sheet_view.showGridLines=False; ws5.sheet_properties.tabColor=C_GREEN
fill_bg(ws5,60,12)
for i in range(1,10): ws5.column_dimensions[get_column_letter(i)].width=22

mth(ws5,1,1,1,9,"CUSTOMER INTELLIGENCE",C_DARK_BG,C_GOLD,20)
mth(ws5,2,1,2,9,"Top customers by revenue and purchase behaviour  •  Apr 11–26 2026",C_DARK_BG,C_GRAY,11)
mth(ws5,4,1,4,7,"  👑  TOP 15 CUSTOMERS BY REVENUE",C_PANEL,C_GREEN,12)

hdr(ws5,5,[(1,"#"),(2,"Customer"),(3,"Revenue (₦)"),(4,"Transactions"),(5,"Units"),(6,"Avg Basket (₦)"),(7,"% of Total")],bgc=C_GREEN,fgc=C_DARK_BG)
for i,(_,r) in enumerate(top_cust.iterrows()):
    row=6+i; alt=i%2==1; avg_basket=r['Revenue']/r['Transactions']; pct=r['Revenue']/total_rev
    drow(ws5,row,[(1,i+1),(2,r['Sold To'].strip()),(3,int(r['Revenue'])),(4,int(r['Transactions'])),(5,int(r['Units'])),(6,int(avg_basket)),(7,f"{pct:.1%}")],alt=alt)
    ws5.cell(row=row,column=3).number_format='#,##0'; ws5.cell(row=row,column=6).number_format='#,##0'
    if i<3: ws5.cell(row=row,column=3).font=ft(C_GOLD,10,bold=True)

ci_row=23; mth(ws5,ci_row,1,ci_row,7,"  💡  CUSTOMER INSIGHTS",C_PANEL,C_GREEN,12)
online_rev=top_cust[top_cust['Sold To'].str.strip()=='TAHIR ONLINE SHOP']['Revenue'].values
online_rev=int(online_rev[0]) if len(online_rev) else 0
top3_pct=top_cust.head(3)['Revenue'].sum()/total_rev
cust_ins=[
    f"• FAHAD TAHIR is #1 at ₦{int(top_cust.iloc[0]['Revenue']):,} ({int(top_cust.iloc[0]['Transactions'])} transactions) — Store Account buyer. Priority VIP service.",
    f"• TOP 3 customers = {top3_pct:.0%} of total revenue — high concentration. Protect these relationships.",
    f"• ABBA TAHIR SHOP: ₦{int(top_cust[top_cust['Sold To'].str.strip()=='Abba Tahir Shop']['Revenue'].values[0]):,} across {int(top_cust[top_cust['Sold To'].str.strip()=='Abba Tahir Shop']['Transactions'].values[0])} transactions — highest frequency buyer.",
    f"• STORE ACCOUNT is the dominant payment channel (39% of transactions) — manage outstanding balances carefully.",
    f"• ONLINE CHANNEL (TAHIR ONLINE SHOP): ₦{online_rev:,} — growing digital revenue. Expand online presence.",
    f"• NEW CUSTOMER MUSA ALARAMMA: Single ₦2.58M transaction — investigate and nurture as VIP.",
]
for i,ins in enumerate(cust_ins):
    row=ci_row+1+i; ws5.row_dimensions[row].height=20
    mth(ws5,row,1,row,7,ins,"1A1A1F" if i%2==0 else C_CARD,C_WHITE,10,False)

# ═══════════════════════ SHEET 6: PERIOD COMPARISON ════════════════════════
ws6 = wb.create_sheet("Period Comparison")
ws6.sheet_view.showGridLines=False; ws6.sheet_properties.tabColor=C_RED
fill_bg(ws6,60,12)
for i in range(1,10): ws6.column_dimensions[get_column_letter(i)].width=20

mth(ws6,1,1,1,8,"PERIOD-OVER-PERIOD COMPARISON",C_DARK_BG,C_GOLD,20)
mth(ws6,2,1,2,8,"31 Mar–11 Apr 2026  vs  11 Apr–26 Apr 2026",C_DARK_BG,C_GRAY,11)
mth(ws6,4,1,4,5,"  📊  KEY METRICS COMPARISON",C_PANEL,C_RED,12)

P1={"Revenue":57098700,"Transactions":525,"Days":11,"AvgDaily":5190791,"AvgBasket":108760,"Units":3377}
P2={"Revenue":total_rev,"Transactions":total_txn,"Days":num_days,"AvgDaily":total_rev/num_days,"AvgBasket":total_rev/total_txn,"Units":total_units}

hdr(ws6,5,[(1,"Metric"),(2,"31 Mar–11 Apr (₦)"),(3,"11 Apr–26 Apr (₦)"),(4,"Change (₦)"),(5,"% Change")],bgc=C_RED)
metrics=[("Total Revenue",P1["Revenue"],P2["Revenue"]),("Transactions",P1["Transactions"],P2["Transactions"]),
         ("Trading Days",P1["Days"],P2["Days"]),("Avg Daily Revenue",P1["AvgDaily"],P2["AvgDaily"]),
         ("Avg Basket Value",P1["AvgBasket"],P2["AvgBasket"]),("Units Sold",P1["Units"],P2["Units"])]
for i,(metric,v1,v2) in enumerate(metrics):
    row=6+i; change=v2-v1; pct=change/v1 if v1 else 0; color=C_GREEN if pct>0 else C_RED
    drow(ws6,row,[(1,metric),(2,v1),(3,v2),(4,change),(5,f"{pct:+.1%}")],alt=i%2==1)
    ws6.cell(row=row,column=5).font=ft(color,10,bold=True)
    for col in [2,3,4]: ws6.cell(row=row,column=col).number_format='#,##0'

mth(ws6,14,1,14,5,"  🧵  CATEGORY COMPARISON",C_PANEL,C_RED,12)
P1_cats={"ATAMPA":45274900,"LACE":6888500,"SHADDA":1732700,"LAFAYYA":1283000,"VOILE":989000,"MATERIAL":548000,"YARD":388100}
hdr(ws6,15,[(1,"Category"),(2,"Period 1 (₦)"),(3,"Period 2 (₦)"),(4,"Change (₦)"),(5,"% Change")],bgc=C_RED)
for i,(_,r) in enumerate(cat_df.iterrows()):
    row=16+i; p1=P1_cats.get(r['Category'],0); p2=int(r['Revenue'])
    change=p2-p1; pct=change/p1 if p1 else 0; color=C_GREEN if pct>0 else C_RED
    drow(ws6,row,[(1,r['Category']),(2,p1),(3,p2),(4,change),(5,f"{pct:+.1%}")],alt=i%2==1)
    ws6.cell(row=row,column=5).font=ft(color,10,bold=True)
    for col in [2,3,4]: ws6.cell(row=row,column=col).number_format='#,##0'

ci_row=25; mth(ws6,ci_row,1,ci_row,8,"  💡  COMPARISON INSIGHTS",C_PANEL,C_GOLD,12)
rev_chg=(P2['Revenue']-P1['Revenue'])/P1['Revenue']
daily_chg=(P2['AvgDaily']-P1['AvgDaily'])/P1['AvgDaily']
basket_chg=(P2['AvgBasket']-P1['AvgBasket'])/P1['AvgBasket']
comp_ins=[
    f"📈 REVENUE: ₦{int(P2['Revenue']):,} vs ₦{P1['Revenue']:,} in Period 1 — {rev_chg:+.1%} more revenue across {P2['Days']} days ({P1['Days']} days in P1).",
    f"📊 DAILY AVERAGE: ₦{int(P2['AvgDaily']):,}/day vs ₦{int(P1['AvgDaily']):,}/day — {daily_chg:+.1%} improvement in daily sales velocity.",
    f"🎯 ATAMPA DOMINANCE: now {cat_df[cat_df['Category']=='ATAMPA']['Revenue'].values[0]/total_rev:.0%} of revenue vs {45274900/57098700:.0%} in P1. Consistent core.",
    f"🆕 NEW PRODUCTS: MEDIUM SUPER (new #1), ELIZABETH EMBELLISHED, VIBES METALLIC — strong new entrants in P2.",
    f"⚡ BASKET VALUE: ₦{int(P2['AvgBasket']):,} vs ₦{P1['AvgBasket']:,} in P1 — {basket_chg:+.1%}. Customers spending {'more' if basket_chg>0 else 'less'} per transaction.",
    f"📦 UNITS SOLD: {int(P2['Units']):,} vs {P1['Units']:,} in P1 — {(P2['Units']-P1['Units'])/P1['Units']:+.1%} — volume has {'grown' if P2['Units']>P1['Units'] else 'contracted'}.",
]
for i,ins in enumerate(comp_ins):
    row=ci_row+1+i; ws6.row_dimensions[row].height=20
    mth(ws6,row,1,row,8,ins,"1A1A1F" if i%2==0 else C_CARD,C_WHITE,10,False)

# wb.save('/home/claude/Textile_Analytics_Apr11_Apr26_2026.xlsx')
wb.save('C:/Users/Tahir General/Downloads/Textile_Analytics_Apr11_Apr26_2026.xlsx')
print("Saved!")
