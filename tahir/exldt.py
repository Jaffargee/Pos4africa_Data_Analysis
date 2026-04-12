
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
import warnings
warnings.filterwarnings('ignore')

# ── Data prep ──────────────────────────────────────────────────────────────
df = pd.read_excel('/mnt/user-data/uploads/Detailed_Sales_Report__3_.xlsx')
df['datetime'] = pd.to_datetime(df['Date'], format='%m/%d/%Y-%I:%M %p', errors='coerce')
df['day_of_week'] = df['datetime'].dt.day_name()
df['date_only'] = df['datetime'].dt.date
df['hour'] = df['datetime'].dt.hour
df = df[df['Category'] != 'Discount']

df['price_bucket'] = pd.cut(df['Selling Price'], 
    bins=[0,5000,10000,20000,30000,50000,100000,500000], 
    labels=['<5k','5-10k','10-20k','20-30k','30-50k','50-100k','>100k'])

dow_order = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']

daily = df.groupby('date_only').agg(
    revenue=('Total','sum'), transactions=('Sale Id','nunique'), items=('Quantity Sold','sum')
).reset_index()
daily['datetime_col'] = pd.to_datetime(daily['date_only'])
daily['dow'] = daily['datetime_col'].dt.day_name()

dow_rev = df.groupby('day_of_week').agg(
    revenue=('Total','sum'), transactions=('Sale Id','nunique'), qty=('Quantity Sold','sum')
).reindex(dow_order).reset_index()

cat_rev = df.groupby('Category').agg(
    revenue=('Total','sum'), qty=('Quantity Sold','sum'), avg_price=('Selling Price','mean')
).sort_values('revenue',ascending=False).reset_index()

cat_dow = df.groupby(['Category','day_of_week'])['Total'].sum().reset_index()
best_day = cat_dow.loc[cat_dow.groupby('Category')['Total'].idxmax()].sort_values('Total',ascending=False)

top_prod = df.groupby(['Name','Category']).agg(
    revenue=('Total','sum'), qty=('Quantity Sold','sum'), 
    transactions=('Sale Id','nunique'), avg_price=('Selling Price','mean')
).sort_values('revenue',ascending=False).reset_index().head(20)

price_bkt = df.groupby('price_bucket', observed=True).agg(
    count=('Item Id','count'), qty=('Quantity Sold','sum'), revenue=('Total','sum')
).reset_index()

cust = df.groupby('Sold To').agg(
    revenue=('Total','sum'), transactions=('Sale Id','nunique')
).sort_values('revenue',ascending=False).reset_index().head(10)

hour_df = df.copy()
hour_df['hb'] = pd.cut(hour_df['hour'], bins=[0,6,11,13,18,21,24],
    labels=['Night(0-6)','Morning(7-11)','Midday(11-1)','Afternoon(1-6)','Evening(6-9)','Late(9+)'])
hour_tbl = hour_df.groupby('hb', observed=True).agg(
    transactions=('Sale Id','nunique'), revenue=('Total','sum')
).reset_index()

post_sale = pd.read_csv('/home/claude/post_to_sale.csv')
post_sale_valid = post_sale[post_sale['days_to_first_sale'].notna()].copy()
post_sale_valid['days_to_first_sale'] = post_sale_valid['days_to_first_sale'].astype(int)
conv_dist = post_sale_valid['days_to_first_sale'].value_counts().sort_index().reset_index()
conv_dist.columns = ['Days','Count']
non_conv = post_sale[post_sale['first_sale_date'].isna()][['product','posted_date']].drop_duplicates()

# ── Style helpers ──────────────────────────────────────────────────────────
C = dict(
    dark='1A1A2E', mid='16213E', accent='E94560', gold='F5A623',
    green='27AE60', teal='0F3460', white='FFFFFF', header='0D1B2A',
    sub='2C3E50', text='2D3436', ltgray='F8F9FA', border='BDC3C7',
    amber='F39C12', red='C0392B', redbg='FDECEA', amberbg='FEF9E7',
    bluebg='EAF4FB', greenbg='EAFBEA'
)

def hf(c): return PatternFill("solid", fgColor=c)
def ft(bold=False, sz=11, color='2D3436', italic=False):
    return Font(name='Calibri', bold=bold, size=sz, color=color, italic=italic)
def al(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr(style='thin', color='BDC3C7'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def mtitle(ws, r, c1, c2, text, bg=C['dark'], fg=C['gold'], sz=16, bold=True):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=text)
    cell.fill = hf(bg); cell.font = ft(bold, sz, fg)
    cell.alignment = al('center','center')
    ws.row_dimensions[r].height = 28 if sz >= 14 else 20

def stitle(ws, r, c1, c2, text):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=text)
    cell.fill = hf(C['teal']); cell.font = ft(True, 12, C['gold'])
    cell.alignment = al('left','center')
    ws.row_dimensions[r].height = 22

def hcell(ws, r, c, v, bg=C['sub'], fg=C['gold']):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = hf(bg); cell.font = ft(True, 10, fg)
    cell.alignment = al('center','center')
    cell.border = bdr(color='444444')

def dcell(ws, r, c, v, bg=C['white'], fg=C['text'], bold=False, align='center', nf=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = hf(bg); cell.font = ft(bold, 10, fg)
    cell.alignment = al(align,'center')
    cell.border = bdr()
    if nf: cell.number_format = nf

def write_tbl(ws, sr, sc, headers, rows, nfmts=None, altbg=C['ltgray']):
    for i, h in enumerate(headers):
        hcell(ws, sr, sc+i, h)
    for ri, row in enumerate(rows):
        bg = C['white'] if ri % 2 == 0 else altbg
        for ci, val in enumerate(row):
            nf = nfmts[ci] if nfmts and ci < len(nfmts) else None
            dcell(ws, sr+1+ri, sc+ci, val, bg=bg, nf=nf)
    return sr + 1 + len(rows)

def kpi_block(ws, row_val, row_lbl, pairs):
    """pairs = [(value, label, color, sc, ec), ...]"""
    for val, label, col, sc, ec in pairs:
        ws.merge_cells(start_row=row_val, start_column=sc, end_row=row_val, end_column=ec)
        ws.merge_cells(start_row=row_lbl, start_column=sc, end_row=row_lbl, end_column=ec)
        c = ws.cell(row=row_val, column=sc, value=val)
        c.fill = hf(col); c.font = ft(True, 16, C['white']); c.alignment = al('center','center')
        d = ws.cell(row=row_lbl, column=sc, value=label)
        d.fill = hf(col); d.font = ft(False, 9, C['white']); d.alignment = al('center','center')
    for r in [row_val, row_lbl]: ws.row_dimensions[r].height = 32 if r == row_val else 20

wb = Workbook()

# ════════════════════════ SHEET 1: EXECUTIVE DASHBOARD ══════════════════════
ws1 = wb.active
ws1.title = "Executive Dashboard"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 2

mtitle(ws1, 1, 2, 20, "TEXTILE BUSINESS — SALES INTELLIGENCE REPORT", C['dark'], C['gold'], 18)
mtitle(ws1, 2, 2, 20, "Period: 31 Mar – 11 Apr 2026  |  POS + Social Analytics", C['mid'], 'AAAAAA', 10, False)
ws1.row_dimensions[3].height = 8

kpi_block(ws1, 4, 5, [
    ("₦57,098,700",   "TOTAL REVENUE (11 DAYS)",  C['accent'],   2, 5),
    ("525",            "TOTAL TRANSACTIONS",        C['teal'],     6, 9),
    ("₦27,144",        "AVG SELLING PRICE",         C['green'],   10, 13),
    ("97%",            "POST→SALE IN ≤3 DAYS",      C['amber'],   14, 17),
    ("2",              "SALES STAFF",               C['sub'],     18, 20),
])

ws1.row_dimensions[6].height = 10

# DOW table
stitle(ws1, 7, 2, 11, "  📅  REVENUE BY DAY OF WEEK")
dow_hdr = ['Day','Revenue (₦)','Transactions','Units Sold']
dow_rows = list(zip(dow_rev['day_of_week'], dow_rev['revenue'], dow_rev['transactions'], dow_rev['qty']))
end_dow = write_tbl(ws1, 8, 2, dow_hdr, dow_rows, nfmts=[None,'#,##0','#,##0','#,##0'])
for i,w in enumerate([16,16,16,14]): ws1.column_dimensions[get_column_letter(2+i)].width = w

# DOW bar chart
ch1 = BarChart(); ch1.type='col'; ch1.title='Revenue by Day of Week'
ch1.style=10; ch1.width=16; ch1.height=11
ch1.add_data(Reference(ws1, min_col=3, min_row=8, max_row=8+7), titles_from_data=True)
ch1.set_categories(Reference(ws1, min_col=2, min_row=9, max_row=8+7))
ws1.add_chart(ch1, "G7")

ws1.row_dimensions[end_dow+1].height = 8

# Category table
stitle(ws1, end_dow+2, 2, 11, "  🧵  REVENUE BY FABRIC CATEGORY")
cat_hdr = ['Category','Revenue (₦)','Units Sold','Avg Price (₦)']
cat_rows = [(r.Category, r.revenue, r.qty, round(r.avg_price,0)) for _,r in cat_rev.iterrows()]
end_cat = write_tbl(ws1, end_dow+3, 2, cat_hdr, cat_rows, nfmts=[None,'#,##0','#,##0','#,##0'])

# Category pie
ch2 = PieChart(); ch2.title='Revenue by Category'; ch2.width=16; ch2.height=11
ch2.add_data(Reference(ws1, min_col=3, min_row=end_dow+3, max_row=end_dow+3+len(cat_rows)), titles_from_data=True)
ch2.set_categories(Reference(ws1, min_col=2, min_row=end_dow+4, max_row=end_dow+3+len(cat_rows)))
ws1.add_chart(ch2, f"G{end_dow+2}")

print("✅ Sheet 1 done")

# ════════════════════════ SHEET 2: TIMING INTELLIGENCE ══════════════════════
ws2 = wb.create_sheet("Timing Intelligence")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions['A'].width = 2

mtitle(ws2, 1, 2, 18, "POSTING TIMING INTELLIGENCE", C['dark'], C['gold'], 16)
mtitle(ws2, 2, 2, 18, "Optimal posting windows for maximum sales velocity", C['mid'], 'AAAAAA', 10, False)
ws2.row_dimensions[3].height = 8

stitle(ws2, 4, 2, 10, "  🕐  SALES BY TIME OF DAY")
hr_hdr = ['Time Bucket','Transactions','Revenue (₦)']
hr_rows = list(zip(hour_tbl['hb'], hour_tbl['transactions'], hour_tbl['revenue']))
end_hr = write_tbl(ws2, 5, 2, hr_hdr, hr_rows, nfmts=[None,'#,##0','#,##0'])
for i,w in enumerate([22,16,18]): ws2.column_dimensions[get_column_letter(2+i)].width = w

ch3 = BarChart(); ch3.type='col'; ch3.title='Revenue by Time of Day'
ch3.style=10; ch3.width=16; ch3.height=10
ch3.add_data(Reference(ws2, min_col=4, min_row=5, max_row=5+len(hr_rows)), titles_from_data=True)
ch3.set_categories(Reference(ws2, min_col=2, min_row=6, max_row=5+len(hr_rows)))
ws2.add_chart(ch3, "F4")

ws2.row_dimensions[end_hr+1].height = 8

stitle(ws2, end_hr+2, 2, 12, "  🏆  BEST DAY TO POST — BY CATEGORY")
bd_hdr = ['Category','Best Day to Post','Revenue on Best Day (₦)']
bd_rows = [(r.Category, r.day_of_week, r.Total) for _,r in best_day.iterrows()]
end_bd = write_tbl(ws2, end_hr+3, 2, bd_hdr, bd_rows, nfmts=[None,None,'#,##0'])
for i,w in enumerate([14,20,22]): ws2.column_dimensions[get_column_letter(2+i)].width = w

ws2.row_dimensions[end_bd+1].height = 8

# Insight box
ir = end_bd + 2
ws2.merge_cells(start_row=ir, start_column=2, end_row=ir+8, end_column=18)
ins = ws2.cell(row=ir, column=2)
ins.value = (
    "💡 TIMING INSIGHTS — Expert Commercial Analysis\n\n"
    "• TUESDAY is your peak day: ₦14.9M revenue = 29% of all weekly sales. Post major arrivals on MONDAY 8–10pm.\n"
    "• AFTERNOON (1–6pm) drives 91% of total revenue. Most buyers browse socials during lunch and afternoons.\n"
    "• LACE sells best on SUNDAY & FRIDAY — likely payday/weekend social shopping. Post Lace on Thursday evening.\n"
    "• SHADDA peaks sharply on MONDAY (₦900k vs ₦29k on Tuesday) — wholesale B2B buyers restocking for the week.\n"
    "• SUNDAY is your weakest day — only 9% of transactions. Avoid posting new products Sunday morning.\n"
    "• RECOMMENDATION: Create a weekly posting calendar. Tuesday = ATAMPA, Friday = LACE, Monday = SHADDA."
)
ins.fill = hf("FFF8E1"); ins.font = ft(False, 10, '4A3800')
ins.alignment = al('left','top', wrap=True); ins.border = bdr(color=C['amber'])
ws2.row_dimensions[ir].height = 130

print("✅ Sheet 2 done")

# ════════════════════════ SHEET 3: PRODUCT PERFORMANCE ══════════════════════
ws3 = wb.create_sheet("Product Performance")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 2

mtitle(ws3, 1, 2, 22, "PRODUCT PERFORMANCE ANALYTICS", C['dark'], C['gold'], 16)
mtitle(ws3, 2, 2, 22, "Revenue, velocity, and margin by product", C['mid'], 'AAAAAA', 10, False)
ws3.row_dimensions[3].height = 8

stitle(ws3, 4, 2, 22, "  🥇  TOP 20 PRODUCTS BY REVENUE")
tp_hdr = ['Product','Category','Revenue (₦)','Units Sold','Transactions','Avg Price (₦)','Rev/Transaction (₦)']
tp_rows = []
for _,r in top_prod.iterrows():
    rpt = r['revenue']/r['transactions'] if r['transactions']>0 else 0
    tp_rows.append([r['Name'], r['Category'], r['revenue'], r['qty'], r['transactions'], round(r['avg_price'],0), round(rpt,0)])
end_tp = write_tbl(ws3, 5, 2, tp_hdr, tp_rows, nfmts=[None,None,'#,##0','#,##0','#,##0','#,##0','#,##0'])
for i,w in enumerate([30,14,14,12,14,14,18]): ws3.column_dimensions[get_column_letter(2+i)].width = w

ws3.conditional_formatting.add(f"D6:D{5+len(tp_rows)}",
    ColorScaleRule(start_type='min', start_color='FFFFFF',
                   mid_type='percentile', mid_value=50, mid_color='FFF3CD',
                   end_type='max', end_color='E94560'))

ch4 = BarChart(); ch4.type='bar'; ch4.title='Top 10 Products — Revenue (₦)'
ch4.style=10; ch4.width=22; ch4.height=14
ch4.add_data(Reference(ws3, min_col=4, min_row=5, max_row=14), titles_from_data=True)
ch4.set_categories(Reference(ws3, min_col=2, min_row=6, max_row=14))
r_chart3 = end_tp + 2
ws3.add_chart(ch4, f"B{r_chart3}")

ws3.row_dimensions[end_tp+1].height = 8

print("✅ Sheet 3 done")

# ════════════════════════ SHEET 4: PRICE SENSITIVITY ════════════════════════
ws4 = wb.create_sheet("Price Sensitivity")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions['A'].width = 2

mtitle(ws4, 1, 2, 14, "PRICE SENSITIVITY ANALYSIS", C['dark'], C['gold'], 16)
mtitle(ws4, 2, 2, 14, "Optimal price bands — volume vs value tradeoff", C['mid'], 'AAAAAA', 10, False)
ws4.row_dimensions[3].height = 8

stitle(ws4, 4, 2, 14, "  📊  SALES BY PRICE BUCKET")
total_rev = int(price_bkt['revenue'].sum())
pb_hdr = ['Price Range','Line Items','Units Sold','Revenue (₦)','% of Revenue']
pb_rows = []
for _,r in price_bkt.iterrows():
    pct = round(r['revenue'] / total_rev * 100, 1)
    pb_rows.append([r['price_bucket'], r['count'], round(r['qty'],0), r['revenue'], pct])
end_pb = write_tbl(ws4, 5, 2, pb_hdr, pb_rows, nfmts=[None,'#,##0','#,##0','#,##0','0.0%'])
for i,w in enumerate([14,14,12,16,14]): ws4.column_dimensions[get_column_letter(2+i)].width = w

# Fix: convert pct to decimal for proper % display
for ri, row in enumerate(pb_rows):
    dcell(ws4, 5+1+ri, 2+4, row[4]/100, nf='0.0%')

ws4.conditional_formatting.add(f"F6:F{5+len(pb_rows)}",
    ColorScaleRule(start_type='min', start_color='FFFFFF', end_type='max', end_color='27AE60'))

ch5 = BarChart(); ch5.type='col'; ch5.title='Revenue by Price Band'
ch5.style=10; ch5.width=16; ch5.height=10
ch5.add_data(Reference(ws4, min_col=5, min_row=5, max_row=5+len(pb_rows)), titles_from_data=True)
ch5.set_categories(Reference(ws4, min_col=2, min_row=6, max_row=5+len(pb_rows)))
ws4.add_chart(ch5, "H4")

ws4.row_dimensions[end_pb+1].height = 8
stitle(ws4, end_pb+2, 2, 18, "  💡  PRICE STRATEGY RECOMMENDATIONS")
ps_hdr = ['Strategy','Price Range','Recommendation']
ps_rows = [
    ['Volume Sweet Spot','₦10,000–₦20,000','43.9% of all line items. Highest velocity. Stock deeply: HOLLANDAD, EMBELLISHED MDT.'],
    ['Revenue Sweet Spot','₦10,000–₦20,000','₦19M revenue — best combined volume+value. Promote actively in both groups.'],
    ['Premium Quick-Sell','₦30,000–₦50,000','₦7.5M / 86 items. CHIGANVY COTTON POPLIN excels here. Repost if unsold >5 days.'],
    ['Ultra Luxury','> ₦100,000','₦10.2M from only 41 items. EXCLUSIVE SWISS LACE drives. Target VIP buyers only.'],
    ['Slow-Mover Risk','> ₦50,000','Items take longer. If >7 days unsold: repost with 5–10% discount or bundle offer.'],
    ['Bulk Incentive','₦5,000–₦10,000','Avg 5.9 units/transaction. Offer 5% off at 10+ yards purchase to increase basket size.'],
]
write_tbl(ws4, end_pb+3, 2, ps_hdr, ps_rows)
for i,w in enumerate([22,18,60]): ws4.column_dimensions[get_column_letter(2+i)].width = w

print("✅ Sheet 4 done")

# ════════════════════════ SHEET 5: POST→SALE TRACKER ════════════════════════
ws5 = wb.create_sheet("Post-to-Sale Tracker")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions['A'].width = 2

mtitle(ws5, 1, 2, 16, "POST → FIRST SALE CONVERSION TRACKER", C['dark'], C['gold'], 16)
mtitle(ws5, 2, 2, 16, "Days from social media post to first recorded POS sale", C['mid'], 'AAAAAA', 10, False)
ws5.row_dimensions[3].height = 8

kpi_block(ws5, 4, 5, [
    ("77 / 98",  "Posts That Converted",    C['green'],  2,  5),
    ("0.8 Days", "Avg Time to First Sale",  C['teal'],   6,  9),
    ("97%",      "Sold Within 3 Days",      C['accent'], 10, 13),
    ("21",       "Non-Converting Posts",    C['amber'],  14, 16),
])

ws5.row_dimensions[6].height = 8
stitle(ws5, 7, 2, 10, "  ⚡  DAYS TO FIRST SALE — DISTRIBUTION")
cd_hdr = ['Days to First Sale','# of Products']
cd_rows = list(zip(conv_dist['Days'], conv_dist['Count']))
end_cd = write_tbl(ws5, 8, 2, cd_hdr, cd_rows)
ws5.column_dimensions['B'].width = 22; ws5.column_dimensions['C'].width = 18

ch6 = BarChart(); ch6.type='col'; ch6.title='Days From Post to First Sale'
ch6.style=10; ch6.width=14; ch6.height=10
ch6.add_data(Reference(ws5, min_col=3, min_row=8, max_row=8+len(cd_rows)), titles_from_data=True)
ch6.set_categories(Reference(ws5, min_col=2, min_row=9, max_row=8+len(cd_rows)))
ws5.add_chart(ch6, "E7")

ws5.row_dimensions[end_cd+1].height = 8
stitle(ws5, end_cd+2, 2, 16, "  📋  ALL POSTS — FULL TRACKER")
all_hdr = ['Product','Post Date','First Sale Date','Days to Sale','≤3 Days?']
all_rows = []
for _,r in post_sale.iterrows():
    days = int(r['days_to_first_sale']) if pd.notna(r['days_to_first_sale']) else 'No Sale'
    conv = '✅ YES' if r['sold_within_3_days'] else ('⏳ Delayed' if pd.notna(r['first_sale_date']) else '❌ No Sale')
    all_rows.append([r['product'], str(r['posted_date']), str(r['first_sale_date']) if pd.notna(r['first_sale_date']) else '—', days, conv])
end_all = write_tbl(ws5, end_cd+3, 2, all_hdr, all_rows)
for i,w in enumerate([32,14,16,14,14]): ws5.column_dimensions[get_column_letter(2+i)].width = w

# Highlight non-converters red
for ri, row in enumerate(all_rows):
    if '❌' in str(row[4]):
        for ci in range(5):
            ws5.cell(row=end_cd+4+ri, column=2+ci).fill = hf('FDECEA')
            ws5.cell(row=end_cd+4+ri, column=2+ci).font = ft(sz=10, color='C0392B')
    elif '⏳' in str(row[4]):
        for ci in range(5):
            ws5.cell(row=end_cd+4+ri, column=2+ci).fill = hf('FEF9E7')

ws5.row_dimensions[end_all+1].height = 8
stitle(ws5, end_all+2, 2, 14, "  ❌  NON-CONVERTING POSTS — ACTION REQUIRED")
nc_hdr = ['Product Posted','Post Date','Recommended Action']
nc_rows = [[r['product'], r['posted_date'], 'Repost with new media + refresh price/caption'] for _,r in non_conv.iterrows()]
write_tbl(ws5, end_all+3, 2, nc_hdr, nc_rows)
for i,w in enumerate([32,14,42]): ws5.column_dimensions[get_column_letter(2+i)].width = w

print("✅ Sheet 5 done")

# ════════════════════════ SHEET 6: CUSTOMER INTELLIGENCE ════════════════════
ws6 = wb.create_sheet("Customer Intelligence")
ws6.sheet_view.showGridLines = False
ws6.column_dimensions['A'].width = 2

mtitle(ws6, 1, 2, 16, "CUSTOMER INTELLIGENCE & REVENUE CONCENTRATION", C['dark'], C['gold'], 16)
mtitle(ws6, 2, 2, 16, "Buyer segmentation, loyalty signals, and payment analysis", C['mid'], 'AAAAAA', 10, False)
ws6.row_dimensions[3].height = 8

stitle(ws6, 4, 2, 14, "  👑  TOP 10 CUSTOMERS BY REVENUE")
cu_hdr = ['Customer','Revenue (₦)','Transactions','Avg Spend/Visit (₦)','Revenue Share']
total_all = int(cust['revenue'].sum())
cu_rows = [(r['Sold To'], r['revenue'], r['transactions'], round(r['revenue']/r['transactions'],0), round(r['revenue']/total_all*100,1)) for _,r in cust.iterrows()]
end_cu = write_tbl(ws6, 5, 2, cu_hdr, cu_rows, nfmts=[None,'#,##0','#,##0','#,##0','0.0%'])

# Fix % col
for ri, row in enumerate(cu_rows):
    dcell(ws6, 5+1+ri, 2+4, row[4]/100, nf='0.0%')

for i,w in enumerate([34,16,16,20,16]): ws6.column_dimensions[get_column_letter(2+i)].width = w
ws6.conditional_formatting.add(f"C6:C{5+len(cu_rows)}",
    ColorScaleRule(start_type='min', start_color='FFFFFF', end_type='max', end_color='E94560'))

# Payment
ws6.row_dimensions[end_cu+1].height = 8
stitle(ws6, end_cu+2, 2, 12, "  💳  PAYMENT METHOD ANALYSIS")
pay_hdr = ['Method','Transactions','Revenue (₦)','Revenue Share','Risk Level']
pay_rows = [
    ['Store Account (Credit)',206,29101300,0.510,'🔴 HIGH — Monitor balances'],
    ['Access Bank Transfer', 151,14986350,0.262,'🟢 LOW — Cleared payments'],
    ['Moniepoint POS',       159,12535050,0.219,'🟢 LOW — Instant settlement'],
    ['Cash',                   9,  476000,0.008,'🟡 MEDIUM — Track carefully'],
]
end_pay = write_tbl(ws6, end_cu+3, 2, pay_hdr, pay_rows, nfmts=[None,'#,##0','#,##0','0.0%',None])
for ri,row in enumerate(pay_rows):
    dcell(ws6, end_cu+4+ri, 2+3, row[3], nf='0.0%')
for i,w in enumerate([24,16,16,16,28]): ws6.column_dimensions[get_column_letter(2+i)].width = w

# Staff
ws6.row_dimensions[end_pay+1].height = 8
stitle(ws6, end_pay+2, 2, 12, "  👔  SALES STAFF PERFORMANCE")
st_hdr = ['Salesperson','Revenue (₦)','Transactions','Revenue Share','Action']
st_rows = [
    ['BB (Primary)', 55343700, 514, 0.969, '⚠️ Single point of failure — cross-train urgently'],
    ['ABBA',          1755000,  11, 0.031, '📚 Currently in training — increase responsibilities'],
]
end_st = write_tbl(ws6, end_pay+3, 2, st_hdr, st_rows, nfmts=[None,'#,##0','#,##0','0.0%',None])
for ri,row in enumerate(st_rows):
    dcell(ws6, end_pay+4+ri, 2+3, row[3], nf='0.0%')
for i,w in enumerate([20,16,16,16,42]): ws6.column_dimensions[get_column_letter(2+i)].width = w

ir = end_st + 2
ws6.merge_cells(start_row=ir, start_column=2, end_row=ir+10, end_column=16)
ins = ws6.cell(row=ir, column=2)
ins.value = (
    "💡 CUSTOMER INTELLIGENCE — KEY FINDINGS\n\n"
    "REVENUE CONCENTRATION RISK:\n"
    "• Top 3 buyers (Fahad Tahir, Abba Tahir Shop, Tahir Indoor) = ₦37.1M = 65% of total revenue in 11 days.\n"
    "• This is extremely concentrated. If any one account stops buying, revenue drops >20% instantly.\n"
    "• ACTION: Recruit 5 new wholesale clients. Target market traders, fashion designers, event planners.\n\n"
    "CREDIT RISK:\n"
    "• 51% of revenue (₦29.1M) is on Store Account credit. This is cash sitting in customers' hands.\n"
    "• ACTION: Implement 30-day payment terms. Generate monthly statements. Charge 2% late fee after 45 days.\n\n"
    "STAFF RISK:\n"
    "• BB handles ₦55.3M (97%) alone. This is unsustainable — illness/absence would halt operations.\n"
    "• ACTION: Train ABBA on full sales process this month. Document all pricing and discount rules."
)
ins.fill = hf('E8F8F2'); ins.font = ft(False, 10, '1A5E3A')
ins.alignment = al('left','top', wrap=True); ins.border = bdr(color=C['green'])
ws6.row_dimensions[ir].height = 180

print("✅ Sheet 6 done")

# ════════════════════════ SHEET 7: STRATEGY & PREDICTIONS ═══════════════════
ws7 = wb.create_sheet("Strategy and Predictions")
ws7.sheet_view.showGridLines = False
ws7.column_dimensions['A'].width = 2

mtitle(ws7, 1, 2, 24, "STRATEGIC RECOMMENDATIONS & PREDICTIVE DECISION RULES", C['dark'], C['gold'], 16)
mtitle(ws7, 2, 2, 24, "Expert economic and commercial analysis — Action plan for growth", C['mid'], 'AAAAAA', 10, False)
ws7.row_dimensions[3].height = 8

stitle(ws7, 4, 2, 24, "  🤖  PREDICTIVE RULES — SIMPLE IF-THEN DECISION LOGIC")
pr_hdr = ['If You See This...','And This...','Then Predict...','Take This Action']
pr_rows = [
    ['ATAMPA product, ₦10–20k','Post on Monday–Tuesday','Sells within 1 day (40% chance same day)','Post immediately. No waiting.'],
    ['LACE product, >₦100k','Post any day','50% sell within 3 days, 50% take 5+ days','Post Thursday. Follow up DM Friday.'],
    ['New product, no history','Price >₦50k','Uncertain — low confidence','Post Tuesday. Monitor 4 days. Repost if no sale.'],
    ['Product unsold for 5+ days','Posted any day','Engagement is low — algorithm deprioritised','Repost with new photo + new time. Try 1pm.'],
    ['HOLLANDAD or EMBELLISHED MDT','Low stock (<20 units)','Stock-out risk within 2–3 days','Reorder immediately. These turn 10 units/txn.'],
    ['SHADDA product, large qty','Monday post','B2B buyers active — likely bulk purchase','Post Monday 7am. Message top 3 SHADDA buyers.'],
    ['Price in ₦30–50k range','Unsold 4+ days','Too expensive for casual buyers','Offer 8% discount or bundle with ₦10k product.'],
    ['Tuesday, Afternoon 1–4pm','Any ATAMPA post','Peak conversion window','Schedule your best new arrivals here.'],
]
end_pr = write_tbl(ws7, 5, 2, pr_hdr, pr_rows)
for i,w in enumerate([30,26,34,36]): ws7.column_dimensions[get_column_letter(2+i)].width = w

ws7.row_dimensions[end_pr+1].height = 8
stitle(ws7, end_pr+2, 2, 24, "  📌  TOP 10 STRATEGIC PRIORITIES (Priority Order)")
sp_hdr = ['Priority','Area','Recommendation','Expected Business Impact']
sp_rows = [
    ['🔴 P1 — CRITICAL','Customer Concentration','Recruit 5+ new wholesale accounts. Top 3 buyers = 65% revenue. Catastrophic if any churns.','Reduce revenue-at-risk by 30%'],
    ['🔴 P2 — CRITICAL','Staff Risk','Cross-train ABBA fully. Document all pricing rules. BB cannot be only seller.','Business continuity protection'],
    ['🔴 P3 — CRITICAL','Credit Control','Set 30-day payment terms. ₦29M outstanding on Store Account. Generate weekly statements.','Improve cash flow ₦10–15M'],
    ['🟠 P4 — HIGH','Posting Calendar','Standardise: ATAMPA on Tue, LACE on Fri/Sun, SHADDA on Mon. Post Mon 8pm for Tue peak.','Estimated +10–15% conversion rate'],
    ['🟠 P5 — HIGH','Non-Converters','21 products with zero sales. Repost with new media and 5–10% price adjustment.','Recover ₦2–5M in dormant stock'],
    ['🟠 P6 — HIGH','Price Architecture','3-tier bundle: Economy ≤₦10k, Mid ₦10–30k, Premium >₦50k. Upsell mid-tier buyers.','Increase avg basket 15%'],
    ['🟡 P7 — MEDIUM','Inventory Depth','Keep HOLLANDAD minimum 50 units always. It turns 10 units/transaction, 51 transactions.','Prevent ₦800k+ stockout losses'],
    ['🟡 P8 — MEDIUM','Post Tracking','Add post_id to POS comment field every sale. Enables proper attribution analysis.','Unlock full analytics capability'],
    ['🟡 P9 — MEDIUM','VIP Loyalty','Priority new-arrival previews for Fahad Tahir & Abba Tahir. 2% cashback at ₦5M.','Lock in top 2 accounts'],
    ['🟢 P10 — GROWTH','ML Readiness','After 90 days of tagged data, build revenue prediction model (Excel or Python).','Automate 30% of repost decisions'],
]
end_sp = write_tbl(ws7, end_pr+3, 2, sp_hdr, sp_rows)
for i,w in enumerate([18,18,60,34]): ws7.column_dimensions[get_column_letter(2+i)].width = w

# Color priority rows
pcolors = {'🔴':C['redbg'],'🟠':C['amberbg'],'🟡':C['bluebg'],'🟢':C['greenbg']}
for ri, row in enumerate(sp_rows):
    icon = row[0][:2]
    bg = pcolors.get(icon, C['white'])
    for ci in range(4):
        ws7.cell(row=end_pr+4+ri, column=2+ci).fill = hf(bg)

ws7.row_dimensions[end_sp+1].height = 8

stitle(ws7, end_sp+2, 2, 24, "  🗺️  90-DAY EXECUTION ROADMAP")
rm_hdr = ['Timeline','Initiative','Owner','Success Metric']
rm_rows = [
    ['Week 1–2','Document all credit accounts. Set limits. Send first statements.','Finance/Owner','100% accounts with defined limits'],
    ['Week 1–2','Cross-train ABBA: shadow BB, handle 10 solo transactions.','Management','ABBA closes ₦500k independently'],
    ['Week 2–3','Repost 21 non-converting products with fresh media + adjusted pricing.','Social/Sales','≥10 products convert within 7 days'],
    ['Week 3–4','Contact 10 potential wholesale buyers. Offer sample parcels.','Sales/Owner','2+ new accounts opened'],
    ['Month 2','Implement post_id tagging in POS Comment field for all sales.','IT/Admin','100% sales tagged by end of month'],
    ['Month 2','Launch Economy / Mid / Premium price bundle tiers.','Commercial','Avg basket size increases 15%'],
    ['Month 2','Build weekly posting calendar. Assign products to days/times.','Marketing','Posting consistency 80%+ weeks'],
    ['Month 3','Analyse 60 days of tagged data. Build first revenue forecast model.','Analytics','Forecast within 20% actual'],
    ['Month 3','Launch VIP loyalty programme for top 5 customers.','Sales','Top 5 customer spend +10%'],
    ['Month 3','Recruit and onboard 3rd salesperson if revenue trajectory holds.','HR/Owner','3rd seller handling 15% of revenue'],
]
write_tbl(ws7, end_sp+3, 2, rm_hdr, rm_rows)
for i,w in enumerate([14,52,18,34]): ws7.column_dimensions[get_column_letter(2+i)].width = w
